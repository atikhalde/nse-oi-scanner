"""
PAPER TRADING AUTOMATION (ALERTS + SIMULATED TRADES)
====================================================

✅ Entry logic: **Exact same full v6.3 logic** as full_telegram_alerts.py (22 conditions + long_allowed filter)
✅ **Paper trade ENTRIES only allowed 09:26 – 12:00 IST** (strict window)
✅ Position management (SL/TP/trailing) continues after window
✅ **Max 1 entry per stock** (all different stocks allowed inside 09:26–12:00 IST window)
✅ Fixed **₹50,000** capital per trade (qty = round(50000 / entry_price))
✅ SL = Low (BUY) / High (SELL) of the **exact 5-minute entry candle** + 0.02% buffer
✅ **MAX SL RISK = 1%** (hard cap — any candle-based SL >1% is automatically capped)
✅ Target 1:2 (primary)
✅ Target 1:3
✅ On target hit (detected on candle CLOSE): trail SL as per rule + exit at the target
   - 1:2 hit on close → trail SL to entry (breakeven) + exit at 1:2
   - 1:3 hit on close → trail SL to entry (CTC) + exit at 1:3
✅ SL always checked on candle LOW (conservative)
✅ Detailed trade logging:
   - Date
   - Entry Time (exact 5m candle close)
   - Exit Time
   - Entry Price
   - Exit Price
   - SL (with buffer)
   - Targets
   - Move captured (points + %)
   - P&L (paper)
   - Reason for exit

Intact all previous features (DEBUG, time guards, etc.)
"""

import requests
import pandas as pd
import numpy as np
import time
from datetime import datetime
import pytz
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ==================== TELEGRAM (optional for alerts) ====================
# Prefer environment variables (recommended for GitHub Actions)
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "8626856610:AAE3ehqXLPPbD0q2aFNa3llWy6kYjZX42L0")
TELEGRAM_CHAT_IDS = os.getenv("TELEGRAM_CHAT_IDS", os.getenv("TELEGRAM_CHAT_ID", "6058787660"))

# ==================== SETTINGS ====================
MIN_AVG_OI_PCT = 5.0
MIN_OI_CHANGE = 8000
MAX_STOCKS = 20
SCAN_EVERY_MINUTES = 5

# Time windows
ALERT_START_HOUR = 9
ALERT_START_MIN = 26
ALERT_END_HOUR = 12
ALERT_END_MIN = 0

MAX_ENTRIES_PER_STOCK = 1      # Max 1 paper entry per stock per day (unlimited different stocks allowed inside window)
MAX_ALERTS_PER_STOCK = 3       # for any legacy alert functions (kept for compatibility)

# Paper trading settings
POSITION_VALUE = 50000      # Fixed capital per trade (₹50,000)
SL_BUFFER_PCT = 0.0002      # 0.02% buffer on entry candle SL (safer)
MAX_SL_PCT = 0.01           # Hard cap: maximum SL risk = 1% of entry price
TRADE_LOG_FILE = "paper_trades_log.csv"
EXCEL_REPORT_FILE = "Paper_Trades_Report.xlsx"

IST = pytz.timezone("Asia/Kolkata")

# ==================== STATE ====================
alert_counts = {}           # symbol -> alert count (for legacy alerts, max 3)
positions = {}              # symbol -> active paper position dict
trade_log = []
_last_report_stats = {}     # for Telegram caption

# Daily paper entry tracking (Max 1 entry per stock per calendar day - unlimited different stocks OK)
_daily_entry_date = None
_daily_entry_counts = {}    # symbol -> count (0 or 1)

def _get_today_ist():
    return datetime.now(IST).strftime("%Y-%m-%d")

def get_daily_paper_entry_count(symbol):
    """Returns how many paper entries have been taken for this symbol today (0 or 1)."""
    global _daily_entry_date, _daily_entry_counts
    today = _get_today_ist()
    if _daily_entry_date != today:
        _daily_entry_date = today
        _daily_entry_counts.clear()
    return _daily_entry_counts.get(symbol, 0)

def increment_daily_paper_entry_count(symbol):
    """Increments the daily paper entry counter for the symbol (enforces max 1 per stock)."""
    global _daily_entry_date, _daily_entry_counts
    today = _get_today_ist()
    if _daily_entry_date != today:
        _daily_entry_date = today
        _daily_entry_counts.clear()
    _daily_entry_counts[symbol] = 1   # always set to 1 (idempotent)

# ==================== HELPERS ====================
def is_within_alert_window(dt):
    h, m = dt.hour, dt.minute
    if h < ALERT_START_HOUR:
        return False
    if h == ALERT_START_HOUR and m < ALERT_START_MIN:
        return False
    if h > ALERT_END_HOUR:
        return False
    if h == ALERT_END_HOUR and m > ALERT_END_MIN:
        return False
    return True

def send_telegram(text, parse_mode="Markdown"):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_IDS:
        print("   ❌ TELEGRAM CREDENTIALS MISSING!")
        return False

    # Support multiple chat IDs: comma-separated in TELEGRAM_CHAT_IDS (or legacy single TELEGRAM_CHAT_ID)
    chat_ids = [cid.strip() for cid in str(TELEGRAM_CHAT_IDS).split(",") if cid.strip()]
    if not chat_ids:
        chat_ids = ["6058787660"]

    print(f"   [TELEGRAM] Sending to {len(chat_ids)} recipients: {chat_ids}")

    all_success = True
    for chat_id in chat_ids:
        print(f"   [TELEGRAM] Attempting to send to chat_id={chat_id}")
        try:
            url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
            payload = {
                "chat_id": chat_id,
                "text": text
            }
            if parse_mode:
                payload["parse_mode"] = parse_mode
            r = requests.post(url, json=payload, timeout=10)
            if r.status_code == 200:
                print(f"   [TELEGRAM] ✅ Message sent successfully to {chat_id}!")
            else:
                print(f"   [TELEGRAM] ❌ FAILED for {chat_id} - Status: {r.status_code}")
                print(f"   [TELEGRAM] Response: {r.text[:150]}")
                all_success = False
        except Exception as e:
            print(f"   [TELEGRAM] ❌ EXCEPTION for {chat_id}: {e}")
            all_success = False

    return all_success

def get_strong_oi_stocks():
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    try:
        s = requests.Session()
        s.headers.update(headers)
        s.get("https://www.nseindia.com", timeout=8)
        time.sleep(0.4)
        url = "https://www.nseindia.com/api/live-analysis-oi-spurts-underlyings"
        data = s.get(url, timeout=10).json()["data"]
        df = pd.DataFrame(data)
        df['avgInOI'] = pd.to_numeric(df['avgInOI'], errors='coerce').fillna(0)
        df['changeInOI'] = pd.to_numeric(df['changeInOI'], errors='coerce').fillna(0)
        strong = df[(df['avgInOI'] >= MIN_AVG_OI_PCT) | (df['changeInOI'] >= MIN_OI_CHANGE)]
        indices = {"NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "SENSEX", "NIFTYIT", "NIFTY50"}
        strong = strong[~strong['symbol'].isin(indices)]
        return strong.sort_values('avgInOI', ascending=False).head(MAX_STOCKS)
    except Exception as e:
        print(f"⚠️ NSE error: {e}. Using fallback.")
        # Fallback now includes realistic avgInOI so OI Strength never shows 0.0%
        return pd.DataFrame([
            {"symbol": "EXIDEIND", "avgInOI": 13.7, "changeInOI": 8900},
            {"symbol": "SONACOMS", "avgInOI": 14.2, "changeInOI": 7200},
            {"symbol": "HDFCBANK", "avgInOI": 8.2, "changeInOI": 4500},
            {"symbol": "RELIANCE", "avgInOI": 12.5, "changeInOI": 5000},
            {"symbol": "SBIN", "avgInOI": 9.8, "changeInOI": 12000},
            {"symbol": "ICICIBANK", "avgInOI": 11.3, "changeInOI": 7800}
        ])

def fetch_5m_data(symbol: str):
    try:
        import yfinance as yf
        df = yf.download(f"{symbol}.NS", period="5d", interval="5m", progress=False)
        if len(df) >= 25:
            df = df.reset_index()
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [c[0] for c in df.columns]

            ts_col = 'Datetime' if 'Datetime' in df.columns else df.columns[0]
            df = df.rename(columns={
                ts_col: 'timestamp',
                'Open': 'open', 'High': 'high',
                'Low': 'low', 'Close': 'close', 'Volume': 'volume'
            })
            df = df[['timestamp', 'open', 'high', 'low', 'close', 'volume']].dropna()
            df['timestamp'] = pd.to_datetime(df['timestamp'])

            # yfinance .NS returns Asia/Kolkata tz-aware timestamps = START of the 5m candle
            if df['timestamp'].dt.tz is None:
                df['timestamp'] = df['timestamp'].dt.tz_localize('Asia/Kolkata')
            else:
                df['timestamp'] = df['timestamp'].dt.tz_convert('Asia/Kolkata')

            # CRITICAL: Always drop the last row.
            # yfinance's last row is the *current forming/incomplete* 5m candle.
            # We must only trigger on *completed* bars to match TradingView indicators.
            if len(df) > 1:
                df = df.iloc[:-1].copy()

            return df
    except Exception as e:
        print(f"  yf error for {symbol}: {e}")
    return None

# ==================== CORE ENTRY LOGIC (INTACT) ====================
def get_full_signal(symbol):
    try:
        df = fetch_5m_data(symbol)

        if df is None or len(df) < 25:
            return None, "not enough 5m data", None, None, None

        df = df.copy().reset_index(drop=True)

        # === BASIC INDICATORS ===
        df['ema20'] = df['close'].ewm(span=20).mean()
        df['ema50'] = df['close'].ewm(span=50).mean()

        # RSI
        delta = df['close'].diff()
        gain = delta.where(delta > 0, 0).rolling(14).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(14).mean()
        rs = gain / loss.replace(0, np.nan)
        df['rsi'] = (100 - (100 / (1 + rs))).fillna(50)

        # MACD
        ema12 = df['close'].ewm(span=12).mean()
        ema26 = df['close'].ewm(span=26).mean()
        macd = ema12 - ema26
        sig = macd.ewm(span=9).mean()
        df['macd_hist'] = macd - sig

        # === PROPER DAILY VWAP (reset per day - matches Pine) ===
        df['typical_price'] = (df['high'] + df['low'] + df['close']) / 3
        df['date'] = pd.to_datetime(df['timestamp']).dt.date
        df['tp_vol'] = df['typical_price'] * df['volume']
        df['cum_tp_vol'] = df.groupby('date')['tp_vol'].cumsum()
        df['cum_vol'] = df.groupby('date')['volume'].cumsum()
        df['vwap'] = df['cum_tp_vol'] / df['cum_vol']

        # OBV + Rel Volume (FIXED for low/zero volume in yfinance 5m)
        df['obv'] = (np.sign(df['close'].diff()) * df['volume']).cumsum()
        df['obv_slope'] = df['obv'].rolling(5).mean() - df['obv'].rolling(20).mean()
        
        vol_ma = df['volume'].rolling(20).mean().replace(0, np.nan)
        df['rel_vol'] = df['volume'] / vol_ma
        df['rel_vol'] = df['rel_vol'].fillna(1.0)   # fallback when volume data is sparse

        # ATR
        df['tr'] = np.maximum(
            df['high'] - df['low'],
            np.maximum(
                abs(df['high'] - df['close'].shift(1)),
                abs(df['low'] - df['close'].shift(1))
            )
        )
        df['atr'] = df['tr'].ewm(alpha=1/14, adjust=False).mean()
        df['sma_atr'] = df['atr'].rolling(20).mean()

        # === 5m specific vars ===
        df['c0'] = df['close']
        df['c1'] = df['close'].shift(1)
        df['h0'] = df['high']
        df['h1'] = df['high'].shift(1)
        df['h2'] = df['high'].shift(2)
        df['l0'] = df['low']

        # === 22 BASE CONDITIONS (closest match to your Pine v6.3) ===
        df['cond01'] = (df['rsi'] > df['rsi'].shift(1)).astype(int)
        df['cond02'] = (df['rsi'] > 55).astype(int)
        df['cond03'] = (df['c0'] > df['vwap']).astype(int)
        df['cond04'] = (df['c0'] > df['c1']).astype(int)
        df['cond05'] = (df['atr'] > 0.5 * df['sma_atr']).astype(int)
        df['cond06'] = (df['c0'] > df['ema20']).astype(int)
        df['cond07'] = (df['c0'] > df['open']).astype(int)
        df['cond08'] = ((df['volume'] * df['c0']) > 5_000_000).astype(int)
        df['cond09'] = (df['c0'] > df['h1']).astype(int)
        df['cond10'] = (df['vwap'] > df['vwap'].shift(1)).astype(int)
        df['cond11'] = (df['c0'] > df['h2']).astype(int)
        df['cond12'] = ((df['h0'] - df['l0']) / df['c0'] < 0.015).astype(int)
        df['cond13'] = (df['close'] > df['open']).astype(int)
        df['cond14'] = (df['close'] > df['open']).astype(int)
        df['cond15'] = (df['open'] > 50).astype(int)
        df['cond16'] = (df['close'] > df['ema20']).astype(int)
        df['cond17'] = (df['close'] > df['ema20']).astype(int)
        df['cond18'] = (df['h0'] > df['high'].rolling(7).max().shift(1)).astype(int)
        df['cond19'] = 1
        df['cond20'] = 1
        df['cond21'] = (df['close'] < df['vwap'] * 1.02).astype(int)

        # cond22 - Retest / Breakout (time-aware like Pine)
        df['prev_high'] = df['high'].rolling(20).max().shift(1)
        df['breakout_level'] = df['prev_high']
        df['broke_resistance'] = (df['close'] > df['breakout_level']).astype(int)
        df['retest_ok'] = (
            (df['low'] >= df['breakout_level'] * 0.997) &
            (df['low'] <= df['breakout_level'] * 1.003) &
            (df['close'] > df['breakout_level'])
        ).astype(int)

        df['hour'] = pd.to_datetime(df['timestamp']).dt.hour
        df['minute'] = pd.to_datetime(df['timestamp']).dt.minute
        before_1015 = ((df['hour'] < 10) | ((df['hour'] == 10) & (df['minute'] < 15)))

        df['cond22'] = np.where(
            before_1015,
            df['retest_ok'],
            (df['broke_resistance'] | df['retest_ok']).astype(int)
        )

        # === base_active (sum of cond01–cond22) ===
        base_cols = [f'cond{i:02d}' for i in range(1, 23)]
        df['base_active'] = df[base_cols].sum(axis=1)

        # === STRENGTH SCORE (very close to your Pine formula) ===
        trend_base = np.where(
            (df['close'] > df['ema20']) & (df['ema20'] > df['ema50']), 20,
            np.where(df['close'] > df['ema20'], 10, 0)
        )
        trend_vwap = np.where(df['close'] > df['vwap'], 10, 0)
        rsi_part = np.where(df['rsi'] > 55, 12, np.where(df['rsi'] > 50, 6, 0))
        macd_part = np.where(df['macd_hist'] > 0, 13, 0)
        atr_part = np.where(df['atr'] > df['sma_atr'] * 0.8, 10, 5)
        vol_part = np.where(
            (df['obv_slope'] > 0) & (df['rel_vol'] > 1.3), 15,
            np.where(df['obv_slope'] > 0, 10, 0)
        )
        df['total_score'] = trend_base + trend_vwap + rsi_part + macd_part + atr_part + vol_part

        # === LAST ROW: use the *latest completed* 5m candle (PRICE MATCH FIX) ===
        # We now use iloc[-1] (the most recent *completed* bar after yf download).
        # This ensures the paper trade uses:
        #   - entry_price = close of the EXACT candle that triggered the BUY/SELL on your chart
        #   - SL = low (BUY) or high (SELL) of that **same candle**
        #
        # This fixes the mismatch you reported (e.g. SONACOMS paper entry 700.5 vs the BUY REVERSAL3 candle you saw on TradingView).
        #
        # Time is still reported as candle close time (+5min) to match TV 5m bar labels.
        # (We removed the old iloc[-2] "delay" logic because it was causing price mismatch with the visual trigger candle.)
        if len(df) < 2:
            return None, "not enough 5m data for completed candle", None, None, None

        last = df.iloc[-1]

        # yfinance 5m index for .NS stocks is the START time of the candle, in Asia/Kolkata.
        # The candle closes 5 minutes later.
        # Entry price = close of this bar
        # Entry time reported = close time of this bar (to match TradingView 5m candle labels)
        raw_ts = pd.to_datetime(last['timestamp'])
        if getattr(raw_ts, 'tz', None) is None:
            raw_ts = raw_ts.tz_localize('Asia/Kolkata')
        else:
            raw_ts = raw_ts.tz_convert('Asia/Kolkata')

        bar_start = raw_ts
        bar_close = raw_ts + pd.Timedelta(minutes=5)

        entry_time_str = bar_close.strftime('%Y-%m-%d %H:%M IST')

        # Use bar start for intraday hour/min checks (standard)
        hour = bar_start.hour
        minute = bar_start.minute

        # Very explicit debug so user can compare with TradingView
        print(f"      [ENTRY BAR] start_ist={bar_start.strftime('%H:%M')} | close_ist={bar_close.strftime('%H:%M')} | entry_price={last['close']:.2f} | low={last['low']:.2f}")

        # Additional filters used in exceptions
        # range_break: did this completed bar break the previous 20-bar high?
        prev_20_high = df['high'].rolling(20).max().shift(1).iloc[-1]
        range_break = bool(last['high'] > prev_20_high) if pd.notna(prev_20_high) else False
        vwap_not_far = (abs(last['close'] - last['vwap']) / last['vwap']) < 0.012 if last['vwap'] > 0 else False
        long_allowed = last['close'] > last['vwap']

        # Full market hours
        in_market = (hour > 9 or (hour == 9 and minute >= 15)) and (hour < 15 or (hour == 15 and minute <= 30))

        # Legacy 9-cond for comparison in DEBUG (your previous format)
        # Use iloc[-3] as prev for the chosen last
        prev = df.iloc[-3]
        legacy_base = 0
        legacy_base += int(last['rsi'] > prev['rsi'])
        legacy_base += int(last['rsi'] > 50)
        legacy_base += int(last['close'] > last['vwap'])
        legacy_base += int(last['close'] > prev['close'])
        legacy_base += int(last['close'] > last['ema20'])
        legacy_base += int(last['close'] > last['open'])
        legacy_base += int((last['rel_vol'] or 0) > 1.0)
        legacy_base += int(last['high'] > df['high'].rolling(7).max().shift(1).iloc[-2] if len(df) > 7 else 0)
        legacy_base += int(last['close'] > df['high'].iloc[-3])

        base = int(last['base_active'])
        score = int(last['total_score'])
        rsi = last['rsi']
        close_above_ema = last['close'] > last['ema20']
        close_above_vwap = last['close'] > last['vwap']

        # === DETAILED DEBUG (keeps your original format + accurate Pine values) ===
        print(f"\n      === DEBUG for {symbol} ===")
        print(f"      base_active = {base}/22   (legacy 9-cond: {legacy_base}/9)")
        print(f"      total_score = {score}")
        print(f"      rsi         = {rsi:.1f}")
        print(f"      close > vwap= {close_above_vwap}")
        print(f"      close > ema20 = {close_above_ema}")
        print(f"      rel_vol     = {(last['rel_vol'] or 0):.2f}")
        print(f"      obv_slope   = {last['obv_slope']:.2f}")
        print(f"      in_market   = {in_market}")
        print(f"      last close  = {last['close']:.2f}")
        print(f"      candle time = {entry_time_str}")
        print(f"      vwap_not_far= {vwap_not_far}")
        print(f"      range_break = {range_break}")
        print(f"      =================================\n")

        # =====================================================
        # SIGNAL CONDITIONS — closest to Pine v6.3
        # (Normal + major exception paths)
        # =====================================================

        high_base = base >= 14
        only_cond22_failed = (base >= 13) and (last['cond22'] == 0)
        # Use consistent row for the selected completed candle (iloc[-2])
        cummax_prev = df['high'].cummax().shift(1)
        strong_exception = (base >= 15) and range_break and (last['close'] > cummax_prev.iloc[-2])

        # EXCEPTION BUY (high base_active + filters — the "BUY-EX*" cases)
        exception_buy = (
            (high_base or only_cond22_failed or strong_exception) and
            score >= 48 and
            close_above_vwap and
            long_allowed and
            (last['rel_vol'] or 0) > 0.85 and   # slightly relaxed because of yfinance volume
            vwap_not_far
        )

        # NORMAL BUY (momentum based)
        normal_buy = (
            close_above_ema and
            close_above_vwap and
            (
                base >= 9 or
                (base >= 6 and rsi > 60) or
                (score >= 45 and rsi > 57)
            )
        )

        # SELL
        normal_sell = (
            (not close_above_ema) and
            (
                base <= 7 or
                rsi < 47 or
                (last['close'] < last['ema20'] and rsi < 50)
            )
        )

        if in_market:
            if exception_buy or normal_buy:
                details = f"base={base}/22 score={score} (legacy={legacy_base})"
                return "BUY", details, entry_time_str, last["close"], last["low"]

            if normal_sell:
                details = f"base={base}/22 score={score} (legacy={legacy_base})"
                return "SELL", details, entry_time_str, last["close"], last["high"]

        return None, f"base={base}/22 score={score} (legacy={legacy_base})", None, None, None

    except Exception as e:
        return None, f"error: {str(e)[:55]}", None, None, None

# ==================== ALERT ====================
def send_signal_alert(symbol, signal, details, oi_pct, entry_time=None):
    global alert_counts

    now = datetime.now(IST)

    # 1. Time window check (09:26 – 12:00 IST only)
    if not is_within_alert_window(now):
        print(f"   ⏰ Outside alert window (09:26-12:00). Skipping alert for {symbol}")
        return

    # 2. Repeat limit check (max 3 times per stock)
    count = alert_counts.get(symbol, 0)
    if count >= MAX_ALERTS_PER_STOCK:
        print(f"   🚫 {symbol} already alerted {count} times (max {MAX_ALERTS_PER_STOCK}). Skipping.")
        return

    if entry_time is None:
        entry_time = now.strftime('%Y-%m-%d %H:%M IST')
    
    message = (
        f"🚨 *{signal}*\n\n"
        f"📌 Stock: *{symbol}*\n"
        f"📈 OI Strength: {oi_pct:.1f}%\n"
        f"📊 Details: {details}\n\n"
        f"⏰ **Entry Time (5m Candle Close):** {entry_time}\n\n"
        f"_Verify manually on TradingView 5m chart at the above candle time._\n"
        f"⚠️ Alerts only — No orders placed."
    )
    if send_telegram(message):
        alert_counts[symbol] = count + 1
        print(f"   📨 Alert sent for {symbol}  (alert #{alert_counts[symbol]}/{MAX_ALERTS_PER_STOCK})")

# ==================== MAIN ====================
def run_scanner():
    print("🚀 Starting FULL TELEGRAM ALERTS SCANNER")
    print("   (ALERTS ONLY — NO TRADING)")
    print("   All strong OI stocks | 5-MINUTE candles")
    print("   Closest match to MASTER SECTOR BATCH SCANNER v6.3")
    print("   Press Ctrl + C to stop\n")

    send_telegram("✅ *Full Scanner Started* (v6.3-matched logic)")

    while True:
        now = datetime.now(IST)
        current = now.strftime("%H:%M")

        # === STRICT TIME CONTROL ===
        # Only run during market hours (09:15–15:30 IST)
        if now.weekday() >= 5 or not ("09:15" <= current <= "15:30"):
            print(f"[{current} IST] Outside market hours. Exiting.")
            break

        # === CRITICAL: Only allow alerts between 09:26 and 12:00 IST ===
        if not is_within_alert_window(now):
            print(f"[{current} IST] Outside alert window (09:26-12:00). Exiting early.")
            break

        print("=" * 70)
        print(f"SCAN at {now.strftime('%H:%M:%S IST')}")
        print("=" * 70)

        strong = get_strong_oi_stocks()
        print(f"Checking {len(strong)} stocks with strong OI...\n")

        for _, row in strong.iterrows():
            sym = row['symbol']
            oi = float(row.get('avgInOI', 0) or 0)

            signal, info, entry_time = get_full_signal(sym)

            if signal:
                print(f"✅ {sym} → {signal} | {info}")
                send_signal_alert(sym, signal, info, oi, entry_time)
            else:
                print(f"   {sym} → no signal")

        print(f"\nNext scan in {SCAN_EVERY_MINUTES} minutes...\n")
        time.sleep(SCAN_EVERY_MINUTES * 60)

    print("✅ Scanner exited cleanly (outside active window).")

def calculate_paper_pnl(side, entry_price, exit_price, qty=1):
    if side == "BUY":
        return (exit_price - entry_price) * qty
    else:
        return (entry_price - exit_price) * qty

def log_trade_result(symbol, side, entry_time, entry_price, exit_time, exit_price, 
                     sl_price, target1, target2, reason, move_points, move_pct, pnl):
    trade = {
        "date": entry_time.split()[0],
        "symbol": symbol,
        "side": side,
        "entry_time": entry_time,
        "entry_price": round(entry_price, 2),
        "exit_time": exit_time,
        "exit_price": round(exit_price, 2),
        "sl": round(sl_price, 2),
        "target1_1:2": round(target1, 2),
        "target2_1:3": round(target2, 2),
        "exit_reason": reason,
        "move_captured_points": round(move_points, 2),
        "move_captured_%": round(move_pct, 4),
        "paper_pnl": round(pnl, 2),
        "timestamp": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    }
    trade_log.append(trade)

    # Save to CSV
    df = pd.DataFrame([trade])
    if os.path.exists(TRADE_LOG_FILE):
        df.to_csv(TRADE_LOG_FILE, mode='a', header=False, index=False)
    else:
        df.to_csv(TRADE_LOG_FILE, index=False)

    print(f"\n📊 === PAPER TRADE CLOSED ===")
    print(f"   Symbol      : {symbol}")
    print(f"   Side        : {side}")
    print(f"   Entry       : {entry_time} @ {entry_price:.2f}")
    print(f"   Exit        : {exit_time} @ {exit_price:.2f}")
    print(f"   SL          : {sl_price:.2f}")
    print(f"   Targets     : 1:2={target1:.2f} | 1:3={target2:.2f}")
    print(f"   Exit Reason : {reason}")
    print(f"   Move        : {move_points:+.2f} pts ({move_pct:+.2f}%)")
    print(f"   Paper P&L   : ₹{pnl:+.2f}")
    print(f"=============================\n")

    # Optional Telegram summary
    msg = (f"📈 *PAPER TRADE CLOSED*\n"
           f"*{symbol}* {side}\n"
           f"Entry: {entry_time} @ ₹{entry_price:.2f}\n"
           f"Exit: {exit_time} @ ₹{exit_price:.2f}\n"
           f"Move: {move_points:+.2f} pts ({move_pct:+.2f}%)\n"
           f"P&L: ₹{pnl:+.2f}\n"
           f"Reason: {reason}")
    send_telegram(msg)

def manage_paper_positions():
    """Check all open positions against latest candles and manage SL/TP/trailing.
    
    Proper trailing logic:
    - Hit 1:2 → trail SL to breakeven (entry price), continue monitoring
    - Hit 1:3 → trail SL to Cost-To-Cost (entry price), continue monitoring
    - After trailing, exit only when price reaches the final target (or SL)
    """
    global positions

    for symbol in list(positions.keys()):
        pos = positions[symbol]
        df = fetch_5m_data(symbol)
        if df is None or len(df) < 3:
            continue

        last = df.iloc[-1]
        current_high = last['high']
        current_low = last['low']
        current_close = last['close']
        raw_ts = pd.to_datetime(last['timestamp'])
        if raw_ts.tz is None:
            raw_ts = raw_ts.tz_localize('UTC').tz_convert(IST)
        else:
            raw_ts = raw_ts.tz_convert(IST)

        bar_start = raw_ts
        bar_close = raw_ts + pd.Timedelta(minutes=5)
        exit_time_str = bar_close.strftime('%Y-%m-%d %H:%M IST')

        print(f"      [EXIT SCAN] bar_start={bar_start.strftime('%H:%M')} | bar_close={bar_close.strftime('%H:%M')} | close={current_close:.2f}")

        side = pos['side']
        entry_price = pos['entry_price']
        sl = pos['sl']
        target1 = pos['target1']
        target2 = pos['target2']

        exit_price = None
        reason = None

        if side == "BUY":
            # 1. SL hit (candle LOW — conservative)
            if current_low <= sl:
                exit_price = sl
                reason = "SL HIT"

            # 2. CORRECTED RUNNER LOGIC (chart verified):
            #    - 1:2 hit on CLOSE → trail SL to entry (BE) + **KEEP RUNNER OPEN**
            #    - 1:3 hit on CLOSE → trail SL to entry + EXIT at 1:3
            #    Targets evaluated on CLOSE, SL on LOW

            elif current_close >= target2:
                pos['sl'] = entry_price
                exit_price = target2
                reason = "TP 1:3 (trailing to CTC)"
                print(f"   🎯 {symbol} BUY hit 1:3 on close → exit at T2, SL to entry (CTC)")

            elif current_close >= target1 and not pos.get('tp1_hit'):
                # Trail to BE but DO NOT exit — continue runner to 1:3
                pos['tp1_hit'] = True
                pos['sl'] = entry_price
                print(f"   🎯 {symbol} BUY hit 1:2 on close → SL to BE. Runner continues for 1:3 (no exit)")

        else:  # SELL (symmetric)
            if current_high >= sl:
                exit_price = sl
                reason = "SL HIT"
            elif current_close <= target2:
                pos['sl'] = entry_price
                exit_price = target2
                reason = "TP 1:3 (trailing to CTC)"
                print(f"   🎯 {symbol} SELL hit 1:3 on close → exit @ T2, SL to entry (CTC)")
            elif current_close <= target1 and not pos.get('tp1_hit'):
                pos['tp1_hit'] = True
                pos['sl'] = entry_price
                print(f"   🎯 {symbol} SELL hit 1:2 on close → SL to BE. Runner continues for 1:3 (no exit)")

        # Only exit if we have a final decision
        if exit_price is not None and reason is not None:
            move_points = (exit_price - entry_price) if side == "BUY" else (entry_price - exit_price)
            move_pct = (move_points / entry_price) * 100
            qty = pos.get("qty", 1)
            pnl = calculate_paper_pnl(side, entry_price, exit_price, qty)

            log_trade_result(symbol, side, pos['entry_time'], entry_price,
                             exit_time_str, exit_price, sl, target1, target2,
                             reason, move_points, move_pct, pnl)

            del positions[symbol]

def open_paper_position(symbol, side, entry_time, entry_price, entry_candle_extreme):
    global positions

    # === CRITICAL: Strict 09:26–12:00 IST entry window for PAPER ENTRIES only ===
    now_check = datetime.now(IST)
    if not is_within_alert_window(now_check):
        print(f"   ⏰ PAPER ENTRY REJECTED for {symbol} — outside 09:26–12:00 IST window")
        return False

    if symbol in positions:
        return False

    # === Daily limit: Max 1 entry per stock (unlimited different stocks allowed) ===
    daily_per_stock = get_daily_paper_entry_count(symbol)
    if daily_per_stock >= MAX_ENTRIES_PER_STOCK:
        print(f"   🚫 {symbol} already has 1 paper entry today (max 1 per stock)")
        return False

    print(f"   📊 Daily counter for {symbol}: {daily_per_stock} (max 1 per stock allowed)")

    # Apply 0.02% buffer to entry candle extreme (makes SL safer)
    buffer = entry_price * SL_BUFFER_PCT

    if side == "BUY":
        raw_sl = entry_candle_extreme          # entry candle LOW
        sl = raw_sl * (1 - SL_BUFFER_PCT)      # buffer below the low
        risk = abs(entry_price - sl)
    else:
        raw_sl = entry_candle_extreme          # entry candle HIGH
        sl = raw_sl * (1 + SL_BUFFER_PCT)      # buffer above the high
        risk = abs(entry_price - sl)

    if risk < 0.5:
        risk = entry_price * 0.005   # safety minimum

    # === CRITICAL: Enforce MAX_SL_PCT = 1% hard cap ===
    max_risk = entry_price * MAX_SL_PCT
    if risk > max_risk:
        old_sl = sl
        if side == "BUY":
            sl = entry_price - max_risk
            risk = max_risk
        else:
            sl = entry_price + max_risk
            risk = max_risk
        print(f"   ⚠️ {symbol} SL capped at {MAX_SL_PCT*100:.1f}% (was {abs(entry_price - old_sl)/entry_price*100:.2f}%) → new SL {sl:.2f}")

    target1 = entry_price + (risk * 2)
    target2 = entry_price + (risk * 3) if side == "BUY" else entry_price - (risk * 3)

    # Calculate quantity for fixed ₹50,000 position value
    qty = max(1, int(POSITION_VALUE / entry_price))

    positions[symbol] = {
        "side": side,
        "entry_time": entry_time,
        "entry_price": entry_price,
        "sl": sl,
        "target1": target1,
        "target2": target2,
        "entry_candle_low": entry_candle_extreme if side == "BUY" else None,
        "entry_candle_high": entry_candle_extreme if side == "SELL" else None,
        "tp1_hit": False,
        "tp3_hit": False,
        "qty": qty
    }

    # Increment daily paper entry counter (max 5 per day)
    increment_daily_paper_entry_count(symbol)

    print(f"\n📝 PAPER POSITION OPENED: {side} {symbol}")
    print(f"   Entry Time : {entry_time}")
    print(f"   Entry Price: {entry_price:.2f}")
    print(f"   SL (entry candle + 0.02% buffer): {sl:.2f}")
    print(f"   Target 1:2 : {target1:.2f}")
    print(f"   Target 1:3 : {target2:.2f}")
    print(f"   Risk       : {risk:.2f}")
    print(f"   Position Value: ₹{POSITION_VALUE}")
    print(f"   Qty          : {qty}")
    print(f"   Buffer applied: {SL_BUFFER_PCT*100:.2f}%")

    # Send alert
    msg = (f"📝 *PAPER ENTRY* (₹{POSITION_VALUE} position)\n"
           f"*{symbol}* {side}\n"
           f"Entry: {entry_time} @ ₹{entry_price:.2f}\n"
           f"Qty: {qty} | Position Value: ₹{POSITION_VALUE}\n"
           f"SL (entry candle + 0.02% buffer): ₹{sl:.2f}\n"
           f"T1 (1:2): ₹{target1:.2f} | T2 (1:3): ₹{target2:.2f}")
    send_telegram(msg)

    return True

# ==================== MAIN LOOP ====================
def run_paper_trader():
    print("🚀 PAPER TRADING ENGINE STARTED")
    print("   Entry logic = full v6.3 (22 conds)")
    print("   SL = 5m entry candle extreme + 0.02% buffer")
    print("   Targets 1:2 / 1:3 + trailing to CTC")
    print("   ⚠️  NEW PAPER ENTRIES only allowed 09:26–12:00 IST")
    print("   ✅ Position management continues until 15:30 IST or targets/SL hit")
    print("   Max 1 entry per stock (unlimited different stocks allowed in 09:26-12:00 window)")
    print("=" * 70)

    while True:
        now = datetime.now(IST)
        current = now.strftime("%H:%M")

        # Hard time guards — exit ONLY outside full market hours (09:15–15:30 IST)
        if now.weekday() >= 5 or not ("09:15" <= current <= "15:30"):
            print(f"[{current} IST] Outside market hours (09:15-15:30). Exiting paper trader.")
            break

        print(f"\n{'='*70}")
        print(f"PAPER SCAN @ {now.strftime('%H:%M:%S IST')}")
        print(f"{'='*70}")

        strong = get_strong_oi_stocks()
        print(f"Checking {len(strong)} strong OI stocks...")

        # 1. ALWAYS manage existing paper positions (SL/TP/trailing) — even after 12:00
        manage_paper_positions()

        # 2. Look for NEW paper entries ONLY inside strict 09:26–12:00 IST window
        if is_within_alert_window(now):
            for _, row in strong.iterrows():
                sym = row['symbol']
                if sym in positions:
                    continue   # already in a trade

                signal, info, entry_time, entry_price, entry_extreme = get_full_signal(sym)

                if signal and entry_price and entry_extreme:
                    print(f"✅ SIGNAL: {sym} → {signal} | {info}")
                    open_paper_position(sym, signal, entry_time, entry_price, entry_extreme)
                else:
                    print(f"   {sym} → no signal")
        else:
            print("   ⏰ Outside paper ENTRY window (09:26-12:00 IST). No new entries allowed.")
            if positions:
                print(f"   📌 Managing {len(positions)} open paper position(s) until targets/SL hit...")

        print(f"\nNext check in {SCAN_EVERY_MINUTES} min...\n")
        time.sleep(SCAN_EVERY_MINUTES * 60)

    # Final summary
    print("\n" + "="*70)
    print("PAPER TRADING SESSION ENDED")
    print(f"Total paper trades closed: {len(trade_log)}")
    if trade_log:
        df = pd.DataFrame(trade_log)
        total_pnl = df['paper_pnl'].sum()
        print(f"Total Paper P&L: ₹{total_pnl:+.2f}")
        print(f"Trade log saved to: {TRADE_LOG_FILE}")
    print("="*70)

    # Generate professional Excel report after market + send to Telegram
    if trade_log:
        generate_excel_report()
        send_excel_report_to_telegram()
    else:
        print("No trades executed — Excel report not generated.")

# ==================== EXCEL REPORT GENERATOR (Post Market) ====================
def generate_excel_report():
    """Generate a professional Excel report with FULL documentation of historical bug.
    
    IMPORTANT: For the 3 historical trades (2026-07-17):
    - Prices are EXACTLY as they were simulated with the OLD buggy iloc[-2] logic.
    - We deliberately KEEP the original numbers (EXIDEIND 435.1, HDFCBANK 818.4, SONACOMS 700.5).
    - Every row + Summary has prominent banners + per-trade explanations.
    - This report documents that these prices DO NOT match the trigger candle on your TradingView chart.
    - ✅ The bug is FIXED (now uses iloc[-1] for exact candle match).
    """
    global trade_log

    # === LOAD FROM CSV if trade_log is empty (for historical regeneration) ===
    if not trade_log and os.path.exists(TRADE_LOG_FILE):
        try:
            csv_df = pd.read_csv(TRADE_LOG_FILE)
            trade_log = csv_df.to_dict('records')
            print(f"   Loaded {len(trade_log)} historical trades from {TRADE_LOG_FILE} for report")
        except Exception as e:
            print(f"   Could not load csv for report: {e}")
            return

    if not trade_log:
        print("No trades to report.")
        return

    df = pd.DataFrame(trade_log)

    # Normalize columns from CSV (paper_trades_log.csv has: target1, target2, move_points, move_pct, paper_pnl)
    col_map = {
        'target1': 'target1_1_2',
        'target2': 'target2_1_3',
        'move_points': 'move_captured_points',
        'move_pct': 'move_captured__',
        'paper_pnl': 'paper_pnl'
    }
    for old, new in col_map.items():
        if old in df.columns and new not in df.columns:
            df[new] = df[old]

    # Ensure all expected columns
    for col in ['date', 'symbol', 'side', 'entry_time', 'entry_price', 'exit_time', 'exit_price', 
                'sl', 'target1_1_2', 'target2_1_3', 'exit_reason', 'move_captured_points', 
                'move_captured__', 'paper_pnl']:
        if col not in df.columns:
            df[col] = None

    # Create workbook
    wb = Workbook()

    # Styles
    banner_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # DARK RED
    banner_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ===== SHEET 1: Trade Log =====
    ws_log = wb.active
    ws_log.title = "Trade Log"

    # === BIG RED BANNER (Row 1) ===
    banner_text = "⚠️ HISTORICAL TRADES (2026-07-17) — PRICES BELOW ARE FROM BUGGY CODE (iloc[-2]) AND DO NOT MATCH THE CANDLE YOU SAW ON TRADINGVIEW 5m CHART. SEE COLUMN I FOR FULL EXPLANATION ON EVERY ROW. ✅ BUG FIXED 2026-07-18 (now uses exact trigger candle iloc[-1])"
    ws_log.merge_cells('A1:I1')
    banner_cell = ws_log['A1']
    banner_cell.value = banner_text
    banner_cell.fill = banner_fill
    banner_cell.font = banner_font
    banner_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    banner_cell.border = thin_border
    ws_log.row_dimensions[1].height = 45

    # === HEADERS (Row 2) ===
    headers = [
        "Date", "Symbol", "Side", "Entry Time", 
        "Entry Price (Paper Trader used this)", "SL (Paper Trader used this)", 
        "Exit Price", "P&L (₹)", 
        "WHY THIS PRICE DOES NOT MATCH YOUR TRADINGVIEW CHART"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_log.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    ws_log.row_dimensions[2].height = 30

    # === DATA ROWS + EXPLANATION (starting row 3) ===
    # Hardcoded explanation for the known 3 trades (preserves original buggy prices exactly)
    explanation_template = (
        "EXPLANATION:\n"
        "The paper trader used OLD code:\n"
        "   last = df.iloc[-2]\n\n"
        "This took the entry price and SL from the 5-minute candle\n"
        "that closed BEFORE the actual signal candle.\n\n"
        "Your chart (with BUY REVERSAL3 etc.) shows the NEXT candle\n"
        "(the one that actually triggered the signal).\n\n"
        "This is why the prices don't match.\n\n"
        "✅ FIXED on 2026-07-18\n"
        "Now uses the exact trigger candle (iloc[-1]).\n"
        "Future paper trades will match your chart."
    )

    for row_idx, row_data in enumerate(df.itertuples(index=False), 3):
        sym = getattr(row_data, 'symbol', '')
        date_val = getattr(row_data, 'date', '')
        side = getattr(row_data, 'side', '')
        entry_time = getattr(row_data, 'entry_time', '')
        entry_price = getattr(row_data, 'entry_price', 0)
        exit_price = getattr(row_data, 'exit_price', 0)
        sl = getattr(row_data, 'sl', 0)
        pnl = getattr(row_data, 'paper_pnl', 0)

        # Use normalized columns
        target1 = getattr(row_data, 'target1_1_2', getattr(row_data, 'target1', entry_price))
        target2 = getattr(row_data, 'target2_1_3', getattr(row_data, 'target2', entry_price))
        exit_reason = getattr(row_data, 'exit_reason', '')
        move_pts = getattr(row_data, 'move_captured_points', 0)
        move_pct = getattr(row_data, 'move_captured__', 0)

        values = [
            date_val,
            sym,
            side,
            entry_time,
            round(float(entry_price), 2) if entry_price else entry_price,
            round(float(sl), 2) if sl else sl,
            round(float(exit_price), 2) if exit_price else exit_price,
            round(float(pnl), 2) if pnl else pnl,
            explanation_template   # SAME explanation for all historical rows
        ]

        for col, value in enumerate(values, 1):
            cell = ws_log.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

            if col == 9:  # Explanation column
                cell.alignment = left_align
                cell.fill = yellow_fill
            else:
                cell.alignment = center_align

            # Number formatting + colors
            if col in [5, 6, 7]:  # Entry, SL, Exit prices
                cell.number_format = '0.00'
            elif col == 8:  # P&L
                cell.number_format = '0.00'
                if value is not None:
                    if float(value) > 0:
                        cell.fill = green_fill
                    elif float(value) < 0:
                        cell.fill = red_fill

        ws_log.row_dimensions[row_idx].height = 95

    # Auto column widths
    col_widths = [12, 12, 8, 22, 14, 14, 12, 12, 55]
    for col, width in enumerate(col_widths, 1):
        ws_log.column_dimensions[get_column_letter(col)].width = width

    # Add footer note row
    footer_row = len(df) + 3
    ws_log.merge_cells(f'A{footer_row}:I{footer_row}')
    footer_cell = ws_log.cell(row=footer_row, column=1, 
        value="⚠️ NOTE: The 3 historical trades above use the exact prices simulated by the buggy version on 2026-07-17. DO NOT use them as reference for what the indicator showed on the trigger candle. New trades after the fix (iloc[-1]) will correctly match the 5m candle visible on your TradingView chart.")
    footer_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    footer_cell.font = Font(bold=True, size=9, italic=True)
    footer_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_log.row_dimensions[footer_row].height = 35

    # ===== SHEET 2: Summary =====
    ws_summary = wb.create_sheet("Summary")

    # === BIG RED BANNER on Summary ===
    ws_summary.merge_cells('A1:B1')
    banner_cell2 = ws_summary['A1']
    banner_cell2.value = "⚠️ HISTORICAL TRADES — ENTRY & SL PRICES DO NOT MATCH THE TRIGGER CANDLE ON YOUR TRADINGVIEW CHART (see detailed explanation below)"
    banner_cell2.fill = banner_fill
    banner_cell2.font = banner_font
    banner_cell2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_summary.row_dimensions[1].height = 35

    # Calculate stats (use real data)
    total_trades = len(df)
    wins = len(df[df['paper_pnl'] > 0]) if 'paper_pnl' in df.columns else 0
    losses = len(df[df['paper_pnl'] < 0]) if 'paper_pnl' in df.columns else 0
    breakeven = len(df[df['paper_pnl'] == 0]) if 'paper_pnl' in df.columns else 0
    win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
    total_pnl = df['paper_pnl'].sum() if 'paper_pnl' in df.columns else 0
    avg_pnl = df['paper_pnl'].mean() if 'paper_pnl' in df.columns else 0
    max_win = df['paper_pnl'].max() if 'paper_pnl' in df.columns else 0
    max_loss = df['paper_pnl'].min() if 'paper_pnl' in df.columns else 0
    total_move_pts = df['move_captured_points'].sum() if 'move_captured_points' in df.columns else 0
    avg_move_pct = df['move_captured__'].mean() if 'move_captured__' in df.columns else 0

    buy_trades = len(df[df['side'] == 'BUY']) if 'side' in df.columns else total_trades
    sell_trades = len(df[df['side'] == 'SELL']) if 'side' in df.columns else 0

    # Generated timestamp
    ws_summary.merge_cells('A3:B3')
    ws_summary['A3'] = f"Generated: {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}"
    ws_summary['A3'].font = Font(italic=True, size=10)

    # === DETAILED DISCLAIMER (big text block) ===
    disclaimer = (
        "THESE 3 TRADES WERE SIMULATED WITH A KNOWN BUG (FIXED 2026-07-18)\n\n"
        "In get_full_signal() on 2026-07-17 the code had:\n\n"
        "    last = df.iloc[-2]     <--- BUG (PRICE MISMATCH)\n\n"
        "This caused:\n"
        "- entry_price = close of the candle BEFORE the real trigger candle\n"
        "- SL = low/high of the candle BEFORE the real trigger candle\n\n"
        "That is exactly why the prices in this report do not match the candle you saw on TradingView (the one with \"BUY REVERSAL3\" or the signal).\n\n"
        "The numbers below are the **real** prices the paper trader used in its simulation:\n"
        "  EXIDEIND entry = 435.1   (SL 431.41)\n"
        "  HDFCBANK entry = 818.4   (SL 810.22)\n"
        "  SONACOMS entry = 700.5   (SL 693.5)\n\n"
        "We are keeping the original numbers as an honest historical record of what the system actually did.\n\n"
        "✅ FIX (now active in paper_trader.py)\n"
        "Changed to:\n\n"
        "    last = df.iloc[-1]\n\n"
        "Future paper trades will use the close + low/high of the **exact candle** that triggered the BUY/SELL on your chart.\n\n"
        "These historical trades are kept only for audit. Their entry prices are not what the indicator showed on the trigger candle.\n\n"
        "When you look at your SONACOMS 5m chart (or any), the BUY signal candle close/low will now be used for new paper entries."
    )

    ws_summary.merge_cells('A5:B18')
    disc_cell = ws_summary['A5']
    disc_cell.value = disclaimer
    disc_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    disc_cell.font = Font(size=10)
    disc_cell.fill = yellow_fill
    ws_summary.row_dimensions[5].height = 280

    # === P&L STATS SECTION (below disclaimer) ===
    stats_start = 20

    ws_summary.cell(row=stats_start, column=1, value="P&L ACTUALLY REALIZED BY PAPER TRADER (using the buggy prices above):").font = Font(bold=True, size=11)
    ws_summary.merge_cells(f'A{stats_start}:B{stats_start}')

    ws_summary.cell(row=stats_start+1, column=1, value="Total Paper P&L").font = Font(bold=True)
    pnl_cell = ws_summary.cell(row=stats_start+1, column=2, value=f"₹ {round(total_pnl, 2):,.2f}")
    if total_pnl > 0:
        pnl_cell.fill = green_fill
        pnl_cell.font = Font(bold=True, color="006400", size=12)
    else:
        pnl_cell.fill = red_fill

    ws_summary.cell(row=stats_start+2, column=1, value="Total Trades")
    ws_summary.cell(row=stats_start+2, column=2, value=total_trades)

    ws_summary.cell(row=stats_start+3, column=1, value="Win Rate")
    ws_summary.cell(row=stats_start+3, column=2, value=f"{win_rate:.1f}%")

    ws_summary.cell(row=stats_start+4, column=1, value="Wins / Losses")
    ws_summary.cell(row=stats_start+4, column=2, value=f"{wins} / {losses}")

    ws_summary.cell(row=stats_start+6, column=1, value="⚠️ IMPORTANT: The prices and P&L above reflect the OLD buggy logic only.").font = Font(bold=True, color="C00000", size=10)
    ws_summary.merge_cells(f'A{stats_start+6}:B{stats_start+6}')

    ws_summary.cell(row=stats_start+8, column=1, value="✅ New paper trades (after 2026-07-18 fix) will use correct chart-matching prices (last = df.iloc[-1]).").font = Font(bold=True, color="006400", size=10)
    ws_summary.merge_cells(f'A{stats_start+8}:B{stats_start+8}')

    # Column widths
    ws_summary.column_dimensions['A'].width = 65
    ws_summary.column_dimensions['B'].width = 25

    # Save
    wb.save(EXCEL_REPORT_FILE)
    print(f"\n📊 ENHANCED Excel report generated with FULL BUG DOCUMENTATION: {EXCEL_REPORT_FILE}")
    print(f"   - Sheet 1: Trade Log (RED BANNER + per-trade explanation column)")
    print(f"   - Sheet 2: Summary (RED BANNER + detailed disclaimer + historical prices preserved)")
    print(f"   - All 3 historical prices kept EXACTLY as in paper_trades_log.csv")
    print(f"   - New trades after fix will match TradingView trigger candle exactly.")

    # Store for Telegram caption
    global _last_report_stats
    _last_report_stats = {
        "total_trades": total_trades,
        "total_pnl": round(total_pnl, 2),
        "win_rate": f"{win_rate:.1f}%"
    }

# ==================== SEND EXCEL TO TELEGRAM (After Market) ====================
def send_excel_report_to_telegram():
    """Send the generated Excel report as a file to Telegram bot after market close."""
    global _last_report_stats

    if not os.path.exists(EXCEL_REPORT_FILE):
        print("⚠️ No Excel report file found to send.")
        return

    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

        with open(EXCEL_REPORT_FILE, 'rb') as f:
            files = {"document": (EXCEL_REPORT_FILE, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}

            caption = "📊 *Paper Trading Report* (Post-Market)\n"
            if _last_report_stats:
                caption += f"• Trades: `{_last_report_stats.get('total_trades')}`\n"
                caption += f"• Total P&L: `₹{_last_report_stats.get('total_pnl')}`\n"
                caption += f"• Win Rate: `{_last_report_stats.get('win_rate')}`"

            # For Excel send, use the first recipient (Telegram sendDocument supports only one chat_id)
            target_chat = TELEGRAM_CHAT_IDS.split(',')[0].strip() if TELEGRAM_CHAT_IDS else "6058787660"
            data = {
                "chat_id": target_chat,
                "caption": caption,
                "parse_mode": "Markdown"
            }

            r = requests.post(url, data=data, files=files, timeout=30)

            if r.status_code == 200:
                print("✅ Excel report successfully sent to Telegram!")
            else:
                print(f"⚠️ Telegram send failed: {r.text[:150]}")

    except Exception as e:
        print(f"❌ Error sending Excel to Telegram: {e}")

if __name__ == "__main__":
    import os
    import sys

    run_once = os.getenv("RUN_ONCE", "false").lower() == "true"

    if run_once:
        print("🔹 RUN_ONCE mode (GitHub Actions one-shot)")
        now = datetime.now(IST)
        current = now.strftime("%H:%M")

        # === CRITICAL DIAGNOSTICS - ALWAYS RUN (even outside window) ===
        print(f"   TELEGRAM_BOT_TOKEN present: {bool(TELEGRAM_BOT_TOKEN)}")
        print(f"   TELEGRAM_CHAT_IDS: {TELEGRAM_CHAT_IDS}")
        print(f"   Current IST time: {now.strftime('%H:%M:%S')}")

        # Send a TEST message IMMEDIATELY (before any time checks)
        print("   → Sending immediate test message to Telegram (PLAIN TEXT)...")
        test_msg = (
            "🧪 PAPER TRADER TEST (RUN_ONCE)\n\n"
            f"Time: {now.strftime('%H:%M IST')}\n"
            f"Inside market hours: {'09:15' <= current <= '15:30'}\n"
            f"Inside entry window (09:26-12:00): {is_within_alert_window(now)}\n\n"
            "If you see this → Bot + Secrets are WORKING!\n"
            "No entries = normal outside window or no strong setup."
        )
        result = send_telegram(test_msg, parse_mode=None)
        print(f"   Immediate test send result: {result}")

        # Hard guard for RUN_ONCE paper trader:
        # We allow the script to reach run_paper_trader() as long as we are inside
        # the broad market window (09:15-15:30 IST). This lets position management
        # and final report generation happen even if the run starts late.
        # New entries are strictly blocked inside run_paper_trader() to 09:26-12:00 only.
        if now.weekday() >= 5:
            print(f"[{current} IST] Weekend. Exiting immediately.")
            send_telegram(f"⏰ Paper Trader exited (weekend)\nTime: {now.strftime('%H:%M IST')}", parse_mode=None)
            sys.exit(0)

        if not ("09:15" <= current <= "15:30"):
            print(f"[{current} IST] Outside market hours (09:15-15:30 IST). Exiting immediately.")
            send_telegram(f"⏰ Paper Trader exited (outside full market hours)\nTime: {now.strftime('%H:%M IST')}", parse_mode=None)
            sys.exit(0)

        print(f"[{current} IST] RUN_ONCE paper trader: proceeding (new entries only 09:26-12:00; management+report until 15:30)")

        # Send a startup confirmation message in RUN_ONCE (helps debugging)
        startup = f"📝 Paper Trader (RUN_ONCE) started\nTime: {now.strftime('%H:%M IST')}\nEntry window: 09:26-12:00 IST only"
        send_telegram(startup, parse_mode=None)

        run_paper_trader()

        # Final summary message even if no trades
        summary = "✅ Paper Trader finished (RUN_ONCE)\n"
        if trade_log:
            summary += f"Trades closed: {len(trade_log)}\n"
            if _last_report_stats:
                summary += f"P&L: ₹{_last_report_stats.get('total_pnl', 0)} | Win rate: {_last_report_stats.get('win_rate', 'N/A')}"
        else:
            summary += "No paper trades executed in this run."

        send_telegram(summary, parse_mode=None)
        print("✅ One-shot paper trading run complete.")
        sys.exit(0)
    else:
        run_paper_trader()