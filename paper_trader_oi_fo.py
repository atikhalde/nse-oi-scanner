"""
PAPER TRADING AUTOMATION (OI-SPURTS + F&O + INDICATOR)
=====================================================

NEW TRIPLE FILTER (in this exact order):

1. NSE OI-spurts check          → get_strong_oi_stocks()
2. F&O securities check         → get_fo_securities()
   (from https://www.nseindia.com/market-data/top-gainers-losers context
    via live-analysis-variations?index=sec_fut + fallback)
3. Indicator logic (v6.3)       → get_full_signal()

All other behavior is 100% intact from original paper_trader.py:
- Full v6.3 22 conditions + exceptions
- Entry only 09:26–12:00 IST
- Max 1 entry per stock per day
- ₹50,000 fixed positions
- SL = entry candle extreme + 0.02% buffer (1% hard cap)
- Trailing (1:2 BE, 1:3 exit)
- Excel report + Telegram (with historical bug docs)
- RUN_ONCE support
- PRICE MATCH FIX (iloc[-1])

"""

import requests
import pandas as pd
import numpy as np
import time
from datetime import datetime
import pytz
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==================== TELEGRAM ====================
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "8626856610:AAE3ehqXLPPbD0q2aFNa3llWy6kYjZX42L0")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "6058787660")

# ==================== SETTINGS ====================
MIN_AVG_OI_PCT = 5.0
MIN_OI_CHANGE = 8000
MAX_STOCKS = 20
SCAN_EVERY_MINUTES = 5

# Time windows
ALERT_START_HOUR = 9
ALERT_START_MIN = 26
ALERT_END_HOUR = 12
ALERT_END_MIN = 0

MAX_ENTRIES_PER_STOCK = 1
MAX_ALERTS_PER_STOCK = 3

# Paper trading
POSITION_VALUE = 50000
SL_BUFFER_PCT = 0.0002
MAX_SL_PCT = 0.01
TRADE_LOG_FILE = "paper_trades_log_oi_fo.csv"
EXCEL_REPORT_FILE = "Paper_Trades_Report_OI_FO.xlsx"

IST = pytz.timezone("Asia/Kolkata")

# ==================== STATE ====================
alert_counts = {}
positions = {}
trade_log = []
_last_report_stats = {}

_daily_entry_date = None
_daily_entry_counts = {}

def _get_today_ist():
    return datetime.now(IST).strftime("%Y-%m-%d")

def get_daily_paper_entry_count(symbol):
    global _daily_entry_date, _daily_entry_counts
    today = _get_today_ist()
    if _daily_entry_date != today:
        _daily_entry_date = today
        _daily_entry_counts.clear()
    return _daily_entry_counts.get(symbol, 0)

def increment_daily_paper_entry_count(symbol):
    global _daily_entry_date, _daily_entry_counts
    today = _get_today_ist()
    if _daily_entry_date != today:
        _daily_entry_date = today
        _daily_entry_counts.clear()
    _daily_entry_counts[symbol] = 1

# ==================== HELPERS ====================
def is_within_alert_window(dt):
    h, m = dt.hour, dt.minute
    if h < ALERT_START_HOUR:
        return False
    if h == ALERT_START_HOUR and m < ALERT_START_MIN:
        return False
    if h > ALERT_END_HOUR:
        return False
    if h == ALERT_END_HOUR and m > ALERT_END_MIN:
        return False
    return True

def send_telegram(text, parse_mode="Markdown"):
    print(f"   [TELEGRAM] Attempting to send to chat_id={TELEGRAM_CHAT_ID}")
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        payload = {"chat_id": TELEGRAM_CHAT_ID, "text": text}
        if parse_mode:
            payload["parse_mode"] = parse_mode
        r = requests.post(url, json=payload, timeout=10)
        if r.status_code == 200:
            print("   [TELEGRAM] ✅ Message sent successfully!")
            return True
        else:
            print(f"   [TELEGRAM] ❌ FAILED - Status: {r.status_code}")
            return False
    except Exception as e:
        print(f"   [TELEGRAM] ❌ EXCEPTION: {e}")
        return False

# ==================== FILTER 1: NSE OI-SPURTS ====================
def get_strong_oi_stocks():
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    try:
        s = requests.Session()
        s.headers.update(headers)
        s.get("https://www.nseindia.com", timeout=8)
        time.sleep(0.4)
        url = "https://www.nseindia.com/api/live-analysis-oi-spurts-underlyings"
        data = s.get(url, timeout=10).json()["data"]
        df = pd.DataFrame(data)
        df['avgInOI'] = pd.to_numeric(df['avgInOI'], errors='coerce').fillna(0)
        df['changeInOI'] = pd.to_numeric(df['changeInOI'], errors='coerce').fillna(0)
        strong = df[(df['avgInOI'] >= MIN_AVG_OI_PCT) | (df['changeInOI'] >= MIN_OI_CHANGE)]
        indices = {"NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "SENSEX", "NIFTYIT", "NIFTY50"}
        strong = strong[~strong['symbol'].isin(indices)]
        return strong.sort_values('avgInOI', ascending=False).head(MAX_STOCKS)
    except Exception as e:
        print(f"⚠️ NSE OI error: {e}. Using fallback.")
        return pd.DataFrame([
            {"symbol": "EXIDEIND"}, {"symbol": "SONACOMS"}, {"symbol": "HDFCBANK"},
            {"symbol": "RELIANCE"}, {"symbol": "SBIN"}, {"symbol": "ICICIBANK"}
        ])

# ==================== FILTER 2: F&O SECURITIES ====================
def get_fo_securities():
    """
    F&O securities filter (second gate).
    Tries live NSE data from top-gainers-losers context (sec_fut / foSec).
    Falls back to a solid list of liquid F&O stocks.
    """
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    try:
        s = requests.Session()
        s.headers.update(headers)
        s.get("https://www.nseindia.com", timeout=8)
        time.sleep(0.5)

        for endpoint in [
            "https://www.nseindia.com/api/live-analysis-variations?index=sec_fut",
            "https://www.nseindia.com/api/live-analysis-variations?index=foSec",
        ]:
            try:
                r = s.get(endpoint, timeout=12)
                if r.status_code == 200:
                    data = r.json()
                    df = pd.DataFrame(data.get("data", []))
                    if not df.empty and "symbol" in df.columns:
                        fo_df = df[["symbol"]].copy()
                        fo_df["symbol"] = fo_df["symbol"].str.upper()
                        print(f"   ✅ F&O filter: {len(fo_df)} securities from NSE ({endpoint.split('=')[-1]})")
                        return fo_df
            except Exception as e:
                print(f"   F&O endpoint {endpoint} failed: {str(e)[:50]}")
                continue
    except Exception as e:
        print(f"⚠️ F&O securities fetch error: {e}")

    # Solid fallback — liquid F&O stocks (covers most active names)
    fallback = [
        "RELIANCE", "HDFCBANK", "ICICIBANK", "INFY", "TCS", "SBIN", "TATAMOTORS",
        "HINDALCO", "LT", "ITC", "AXISBANK", "BAJFINANCE", "MARUTI", "SUNPHARMA",
        "TATASTEEL", "WIPRO", "HCLTECH", "POWERGRID", "NTPC", "COALINDIA", "ONGC",
        "BPCL", "CIPLA", "DRREDDY", "EICHERMOT", "GRASIM", "HEROMOTOCO", "HINDUNILVR",
        "JSWSTEEL", "M&M", "SHREECEM", "TATACONSUM", "ULTRACEMCO", "UPL", "ADANIENT",
        "ADANIPORTS", "BAJAJFINSV", "BAJAJ-AUTO", "BRITANNIA", "DIVISLAB", "HDFCLIFE",
        "ICICIPRULI", "INDUSINDBK", "KOTAKBANK", "LTIM", "PIDILITIND", "SBILIFE",
        "TATAPOWER", "TECHM", "TITAN", "ZOMATO"
    ]
    print(f"   ⚠️ Using F&O fallback list ({len(fallback)} stocks)")
    return pd.DataFrame([{"symbol": s} for s in fallback])

# ==================== COMBINED FILTER (OI + F&O) ====================
def get_oi_and_fo_stocks():
    """
    Returns stocks that pass BOTH:
    1. NSE OI-spurts (strong OI)
    2. F&O securities list
    """
    oi_df = get_strong_oi_stocks()
    fo_df = get_fo_securities()

    oi_symbols = set(oi_df['symbol'].str.upper())
    fo_symbols = set(fo_df['symbol'].str.upper())

    common = oi_symbols & fo_symbols
    if not common:
        print("   ⚠️ No overlap between OI-spurts and F&O. Using OI list only.")
        return oi_df

    filtered = oi_df[oi_df['symbol'].str.upper().isin(common)].copy()
    print(f"   ✅ Triple-filter ready: {len(filtered)} stocks passed OI + F&O")
    return filtered

# ==================== FETCH 5M DATA ====================
def fetch_5m_data(symbol: str):
    try:
        import yfinance as yf
        df = yf.download(f"{symbol}.NS", period="5d", interval="5m", progress=False)
        if len(df) >= 25:
            df = df.reset_index()
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [c[0] for c in df.columns]
            df = df.rename(columns={
                'Datetime': 'timestamp', 'Open': 'open', 'High': 'high',
                'Low': 'low', 'Close': 'close', 'Volume': 'volume'
            })
            df = df[['timestamp', 'open', 'high', 'low', 'close', 'volume']].dropna()
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            return df
    except Exception as e:
        print(f"  yf error for {symbol}: {e}")
    return None

# ==================== CORE INDICATOR LOGIC (UNCHANGED v6.3) ====================
def get_full_signal(symbol):
    try:
        df = fetch_5m_data(symbol)

        if df is None or len(df) < 25:
            return None, "not enough 5m data", None, None, None

        df = df.copy().reset_index(drop=True)

        # === BASIC INDICATORS ===
        df['ema20'] = df['close'].ewm(span=20).mean()
        df['ema50'] = df['close'].ewm(span=50).mean()

        delta = df['close'].diff()
        gain = delta.where(delta > 0, 0).rolling(14).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(14).mean()
        rs = gain / loss.replace(0, np.nan)
        df['rsi'] = (100 - (100 / (1 + rs))).fillna(50)

        ema12 = df['close'].ewm(span=12).mean()
        ema26 = df['close'].ewm(span=26).mean()
        macd = ema12 - ema26
        sig = macd.ewm(span=9).mean()
        df['macd_hist'] = macd - sig

        # VWAP
        df['typical_price'] = (df['high'] + df['low'] + df['close']) / 3
        df['date'] = pd.to_datetime(df['timestamp']).dt.date
        df['tp_vol'] = df['typical_price'] * df['volume']
        df['cum_tp_vol'] = df.groupby('date')['tp_vol'].cumsum()
        df['cum_vol'] = df.groupby('date')['volume'].cumsum()
        df['vwap'] = df['cum_tp_vol'] / df['cum_vol']

        # OBV + Rel Volume
        df['obv'] = (np.sign(df['close'].diff()) * df['volume']).cumsum()
        df['obv_slope'] = df['obv'].rolling(5).mean() - df['obv'].rolling(20).mean()
        vol_ma = df['volume'].rolling(20).mean().replace(0, np.nan)
        df['rel_vol'] = df['volume'] / vol_ma
        df['rel_vol'] = df['rel_vol'].fillna(1.0)

        # ATR
        df['tr'] = np.maximum(
            df['high'] - df['low'],
            np.maximum(abs(df['high'] - df['close'].shift(1)), abs(df['low'] - df['close'].shift(1)))
        )
        df['atr'] = df['tr'].ewm(alpha=1/14, adjust=False).mean()
        df['sma_atr'] = df['atr'].rolling(20).mean()

        # 5m vars
        df['c0'] = df['close']
        df['c1'] = df['close'].shift(1)
        df['h0'] = df['high']
        df['h1'] = df['high'].shift(1)
        df['h2'] = df['high'].shift(2)
        df['l0'] = df['low']

        # 22 CONDITIONS (identical to original)
        df['cond01'] = (df['rsi'] > df['rsi'].shift(1)).astype(int)
        df['cond02'] = (df['rsi'] > 55).astype(int)
        df['cond03'] = (df['c0'] > df['vwap']).astype(int)
        df['cond04'] = (df['c0'] > df['c1']).astype(int)
        df['cond05'] = (df['atr'] > 0.5 * df['sma_atr']).astype(int)
        df['cond06'] = (df['c0'] > df['ema20']).astype(int)
        df['cond07'] = (df['c0'] > df['open']).astype(int)
        df['cond08'] = ((df['volume'] * df['c0']) > 5_000_000).astype(int)
        df['cond09'] = (df['c0'] > df['h1']).astype(int)
        df['cond10'] = (df['vwap'] > df['vwap'].shift(1)).astype(int)
        df['cond11'] = (df['c0'] > df['h2']).astype(int)
        df['cond12'] = ((df['h0'] - df['l0']) / df['c0'] < 0.015).astype(int)
        df['cond13'] = (df['close'] > df['open']).astype(int)
        df['cond14'] = (df['close'] > df['open']).astype(int)
        df['cond15'] = (df['open'] > 50).astype(int)
        df['cond16'] = (df['close'] > df['ema20']).astype(int)
        df['cond17'] = (df['close'] > df['ema20']).astype(int)
        df['cond18'] = (df['h0'] > df['high'].rolling(7).max().shift(1)).astype(int)
        df['cond19'] = 1
        df['cond20'] = 1
        df['cond21'] = (df['close'] < df['vwap'] * 1.02).astype(int)

        df['prev_high'] = df['high'].rolling(20).max().shift(1)
        df['breakout_level'] = df['prev_high']
        df['broke_resistance'] = (df['close'] > df['breakout_level']).astype(int)
        df['retest_ok'] = (
            (df['low'] >= df['breakout_level'] * 0.997) &
            (df['low'] <= df['breakout_level'] * 1.003) &
            (df['close'] > df['breakout_level'])
        ).astype(int)

        df['hour'] = pd.to_datetime(df['timestamp']).dt.hour
        df['minute'] = pd.to_datetime(df['timestamp']).dt.minute
        before_1015 = ((df['hour'] < 10) | ((df['hour'] == 10) & (df['minute'] < 15)))
        df['cond22'] = np.where(before_1015, df['retest_ok'], (df['broke_resistance'] | df['retest_ok']).astype(int))

        base_cols = [f'cond{i:02d}' for i in range(1, 23)]
        df['base_active'] = df[base_cols].sum(axis=1)

        trend_base = np.where(
            (df['close'] > df['ema20']) & (df['ema20'] > df['ema50']), 20,
            np.where(df['close'] > df['ema20'], 10, 0)
        )
        trend_vwap = np.where(df['close'] > df['vwap'], 10, 0)
        rsi_part = np.where(df['rsi'] > 55, 12, np.where(df['rsi'] > 50, 6, 0))
        macd_part = np.where(df['macd_hist'] > 0, 13, 0)
        atr_part = np.where(df['atr'] > df['sma_atr'] * 0.8, 10, 5)
        vol_part = np.where(
            (df['obv_slope'] > 0) & (df['rel_vol'] > 1.3), 15,
            np.where(df['obv_slope'] > 0, 10, 0)
        )
        df['total_score'] = trend_base + trend_vwap + rsi_part + macd_part + atr_part + vol_part

        # PRICE MATCH FIX - iloc[-1]
        if len(df) < 2:
            return None, "not enough 5m data for completed candle", None, None, None

        last = df.iloc[-1]

        ts = pd.to_datetime(last['timestamp'])
        if ts.tz is None:
            ts = ts.tz_localize('UTC').tz_convert(IST)
        else:
            ts = ts.tz_convert(IST)
        close_time = ts + pd.Timedelta(minutes=5)
        entry_time_str = close_time.strftime('%Y-%m-%d %H:%M IST')
        hour = ts.hour
        minute = ts.minute

        range_break = bool(last['high'] > df['high'].rolling(20).max().shift(1).iloc[-2])
        vwap_not_far = (abs(last['close'] - last['vwap']) / last['vwap']) < 0.012 if last['vwap'] > 0 else False
        long_allowed = last['close'] > last['vwap']

        in_market = (hour > 9 or (hour == 9 and minute >= 15)) and (hour < 15 or (hour == 15 and minute <= 30))

        prev = df.iloc[-3]
        legacy_base = 0
        legacy_base += int(last['rsi'] > prev['rsi'])
        legacy_base += int(last['rsi'] > 50)
        legacy_base += int(last['close'] > last['vwap'])
        legacy_base += int(last['close'] > prev['close'])
        legacy_base += int(last['close'] > last['ema20'])
        legacy_base += int(last['close'] > last['open'])
        legacy_base += int((last['rel_vol'] or 0) > 1.0)
        legacy_base += int(last['high'] > df['high'].rolling(7).max().shift(1).iloc[-2] if len(df) > 7 else 0)
        legacy_base += int(last['close'] > df['high'].iloc[-3])

        base = int(last['base_active'])
        score = int(last['total_score'])
        rsi = last['rsi']
        close_above_ema = last['close'] > last['ema20']
        close_above_vwap = last['close'] > last['vwap']

        print(f"\n      === DEBUG for {symbol} ===")
        print(f"      base_active = {base}/22   (legacy 9-cond: {legacy_base}/9)")
        print(f"      total_score = {score}")
        print(f"      rsi         = {rsi:.1f}")
        print(f"      close > vwap= {close_above_vwap}")
        print(f"      close > ema20 = {close_above_ema}")
        print(f"      rel_vol     = {(last['rel_vol'] or 0):.2f}")
        print(f"      in_market   = {in_market}")
        print(f"      last close  = {last['close']:.2f}")
        print(f"      candle time = {entry_time_str}")
        print(f"      =================================\n")

        high_base = base >= 14
        only_cond22_failed = (base >= 13) and (last['cond22'] == 0)
        cummax_prev = df['high'].cummax().shift(1)
        strong_exception = (base >= 15) and range_break and (last['close'] > cummax_prev.iloc[-2])

        exception_buy = (
            (high_base or only_cond22_failed or strong_exception) and
            score >= 48 and
            close_above_vwap and
            long_allowed and
            (last['rel_vol'] or 0) > 0.85 and
            vwap_not_far
        )

        normal_buy = (
            close_above_ema and
            close_above_vwap and
            (base >= 9 or (base >= 6 and rsi > 60) or (score >= 45 and rsi > 57))
        )

        normal_sell = (
            (not close_above_ema) and
            (base <= 7 or rsi < 47 or (last['close'] < last['ema20'] and rsi < 50))
        )

        if in_market:
            if exception_buy or normal_buy:
                details = f"base={base}/22 score={score} (legacy={legacy_base})"
                return "BUY", details, entry_time_str, last["close"], last["low"]

            if normal_sell:
                details = f"base={base}/22 score={score} (legacy={legacy_base})"
                return "SELL", details, entry_time_str, last["close"], last["high"]

        return None, f"base={base}/22 score={score} (legacy={legacy_base})", None, None, None

    except Exception as e:
        return None, f"error: {str(e)[:55]}", None, None, None

# ==================== ALERT ====================
def send_signal_alert(symbol, signal, details, oi_pct, entry_time=None):
    global alert_counts
    now = datetime.now(IST)

    if not is_within_alert_window(now):
        print(f"   ⏰ Outside alert window (09:26-12:00). Skipping alert for {symbol}")
        return

    count = alert_counts.get(symbol, 0)
    if count >= MAX_ALERTS_PER_STOCK:
        print(f"   🚫 {symbol} already alerted {count} times. Skipping.")
        return

    if entry_time is None:
        entry_time = now.strftime('%Y-%m-%d %H:%M IST')

    message = (
        f"🚨 *{signal}* (OI + F&O + INDICATOR)\n\n"
        f"📌 Stock: *{symbol}*\\n"
        f"📈 OI Strength: {oi_pct:.1f}%\\n"
        f"📊 Details: {details}\\n\\n"
        f"⏰ **Entry Time (5m Candle Close):** {entry_time}\\n\\n"
        f"_Passed: OI-spurts → F&O securities → Indicator_"
    )
    if send_telegram(message):
        alert_counts[symbol] = count + 1
        print(f"   📨 Alert sent for {symbol} (alert #{alert_counts[symbol]}/{MAX_ALERTS_PER_STOCK})")

# ==================== PAPER TRADING FUNCTIONS (INTACT) ====================
def calculate_paper_pnl(side, entry_price, exit_price, qty=1):
    if side == "BUY":
        return (exit_price - entry_price) * qty
    else:
        return (entry_price - exit_price) * qty

def log_trade_result(symbol, side, entry_time, entry_price, exit_time, exit_price,
                     sl_price, target1, target2, reason, move_points, move_pct, pnl):
    trade = {
        "date": entry_time.split()[0],
        "symbol": symbol,
        "side": side,
        "entry_time": entry_time,
        "entry_price": round(entry_price, 2),
        "exit_time": exit_time,
        "exit_price": round(exit_price, 2),
        "sl": round(sl_price, 2),
        "target1_1:2": round(target1, 2),
        "target2_1:3": round(target2, 2),
        "exit_reason": reason,
        "move_captured_points": round(move_points, 2),
        "move_captured_%": round(move_pct, 4),
        "paper_pnl": round(pnl, 2),
        "timestamp": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    }
    trade_log.append(trade)

    df = pd.DataFrame([trade])
    if os.path.exists(TRADE_LOG_FILE):
        df.to_csv(TRADE_LOG_FILE, mode='a', header=False, index=False)
    else:
        df.to_csv(TRADE_LOG_FILE, index=False)

    print(f"\n📊 === PAPER TRADE CLOSED (OI+F&O) ===")
    print(f"   Symbol      : {symbol}")
    print(f"   Side        : {side}")
    print(f"   Entry       : {entry_time} @ {entry_price:.2f}")
    print(f"   Exit        : {exit_time} @ {exit_price:.2f}")
    print(f"   P&L         : ₹{pnl:+.2f}")
    print("=====================================\n")

    msg = (f"📈 *PAPER TRADE CLOSED (OI+F&O)*\\n"
           f"*{symbol}* {side}\\n"
           f"Entry: {entry_time} @ ₹{entry_price:.2f}\\n"
           f"Exit: {exit_time} @ ₹{exit_price:.2f}\\n"
           f"P&L: ₹{pnl:+.2f}\\nReason: {reason}")
    send_telegram(msg)

def manage_paper_positions():
    global positions
    for symbol in list(positions.keys()):
        pos = positions[symbol]
        df = fetch_5m_data(symbol)
        if df is None or len(df) < 3:
            continue

        last = df.iloc[-1]
        current_high = last['high']
        current_low = last['low']
        current_close = last['close']
        current_time = pd.to_datetime(last['timestamp'])
        if current_time.tz is None:
            current_time = current_time.tz_localize('UTC').tz_convert(IST)
        else:
            current_time = current_time.tz_convert(IST)
        exit_time_str = current_time.strftime('%Y-%m-%d %H:%M IST')

        side = pos['side']
        entry_price = pos['entry_price']
        sl = pos['sl']
        target1 = pos['target1']
        target2 = pos['target2']

        exit_price = None
        reason = None

        if side == "BUY":
            if current_low <= sl:
                exit_price = sl
                reason = "SL HIT"
            elif current_close >= target2:
                pos['sl'] = entry_price
                exit_price = target2
                reason = "TP 1:3 (trailing to CTC)"
            elif current_close >= target1 and not pos.get('tp1_hit'):
                pos['tp1_hit'] = True
                pos['sl'] = entry_price
                print(f"   🎯 {symbol} BUY hit 1:2 → SL to BE (runner continues)")
        else:
            if current_high >= sl:
                exit_price = sl
                reason = "SL HIT"
            elif current_close <= target2:
                pos['sl'] = entry_price
                exit_price = target2
                reason = "TP 1:3 (trailing to CTC)"
            elif current_close <= target1 and not pos.get('tp1_hit'):
                pos['tp1_hit'] = True
                pos['sl'] = entry_price

        if exit_price is not None and reason is not None:
            move_points = (exit_price - entry_price) if side == "BUY" else (entry_price - exit_price)
            move_pct = (move_points / entry_price) * 100
            qty = pos.get("qty", 1)
            pnl = calculate_paper_pnl(side, entry_price, exit_price, qty)

            log_trade_result(symbol, side, pos['entry_time'], entry_price,
                             exit_time_str, exit_price, sl, target1, target2,
                             reason, move_points, move_pct, pnl)
            del positions[symbol]

def open_paper_position(symbol, side, entry_time, entry_price, entry_candle_extreme):
    global positions

    now_check = datetime.now(IST)
    if not is_within_alert_window(now_check):
        print(f"   ⏰ PAPER ENTRY REJECTED for {symbol} — outside 09:26–12:00 IST window")
        return False

    if symbol in positions:
        return False

    daily_per_stock = get_daily_paper_entry_count(symbol)
    if daily_per_stock >= MAX_ENTRIES_PER_STOCK:
        print(f"   🚫 {symbol} already has 1 paper entry today")
        return False

    buffer = entry_price * SL_BUFFER_PCT

    if side == "BUY":
        raw_sl = entry_candle_extreme
        sl = raw_sl * (1 - SL_BUFFER_PCT)
        risk = abs(entry_price - sl)
    else:
        raw_sl = entry_candle_extreme
        sl = raw_sl * (1 + SL_BUFFER_PCT)
        risk = abs(entry_price - sl)

    if risk < 0.5:
        risk = entry_price * 0.005

    max_risk = entry_price * MAX_SL_PCT
    if risk > max_risk:
        if side == "BUY":
            sl = entry_price - max_risk
        else:
            sl = entry_price + max_risk
        risk = max_risk
        print(f"   ⚠️ {symbol} SL capped at 1%")

    target1 = entry_price + (risk * 2)
    target2 = entry_price + (risk * 3) if side == "BUY" else entry_price - (risk * 3)
    qty = max(1, int(POSITION_VALUE / entry_price))

    positions[symbol] = {
        "side": side,
        "entry_time": entry_time,
        "entry_price": entry_price,
        "sl": sl,
        "target1": target1,
        "target2": target2,
        "tp1_hit": False,
        "qty": qty
    }

    increment_daily_paper_entry_count(symbol)

    print(f"\n📝 PAPER POSITION OPENED (OI+F&O): {side} {symbol}")
    print(f"   Entry: {entry_time} @ {entry_price:.2f}")
    print(f"   SL: {sl:.2f} | T1: {target1:.2f} | T2: {target2:.2f}")

    msg = (f"📝 *PAPER ENTRY (OI + F&O)*\\n"
           f"*{symbol}* {side}\\n"
           f"Entry: {entry_time} @ ₹{entry_price:.2f}\\n"
           f"SL: ₹{sl:.2f} | T1: ₹{target1:.2f} | T2: ₹{target2:.2f}")
    send_telegram(msg)
    return True

# ==================== MAIN LOOP ====================
def run_paper_trader():
    print("🚀 PAPER TRADER (OI-SPURTS + F&O + INDICATOR) STARTED")
    print("   Filter order: 1. OI-spurts → 2. F&O securities → 3. Indicator (v6.3)")
    print("   All other logic intact (trailing, Excel, etc.)")
    print("=" * 70)

    while True:
        now = datetime.now(IST)
        current = now.strftime("%H:%M")

        if now.weekday() >= 5 or not ("09:15" <= current <= "15:30"):
            print(f"[{current} IST] Outside market hours. Exiting.")
            break

        print(f"\n{'='*70}")
        print(f"PAPER SCAN (OI+F&O) @ {now.strftime('%H:%M:%S IST')}")
        print(f"{'='*70}")

        # === TRIPLE FILTER ===
        candidates = get_oi_and_fo_stocks()
        print(f"Checking {len(candidates)} stocks that passed OI + F&O...")

        # 1. Manage existing positions (always)
        manage_paper_positions()

        # 2. New entries only in window
        if is_within_alert_window(now):
            for _, row in candidates.iterrows():
                sym = row['symbol']
                if sym in positions:
                    continue

                signal, info, entry_time, entry_price, entry_extreme = get_full_signal(sym)

                if signal and entry_price and entry_extreme:
                    print(f"✅ SIGNAL (OI+F&O+IND): {sym} → {signal} | {info}")
                    open_paper_position(sym, signal, entry_time, entry_price, entry_extreme)
                else:
                    print(f"   {sym} → no indicator signal")
        else:
            print("   ⏰ Outside paper ENTRY window (09:26-12:00 IST). No new entries.")

        print(f"\nNext check in {SCAN_EVERY_MINUTES} min...\n")
        time.sleep(SCAN_EVERY_MINUTES * 60)

    print("\n" + "="*70)
    print("PAPER TRADING SESSION ENDED")
    if trade_log:
        df = pd.DataFrame(trade_log)
        print(f"Total paper trades closed: {len(trade_log)}")
        print(f"Total Paper P&L: ₹{df['paper_pnl'].sum():+.2f}")
        generate_excel_report()
        send_excel_report_to_telegram()
    else:
        print("No trades executed.")
    print("="*70)

# ==================== EXCEL REPORT (identical logic, new filename) ====================
def generate_excel_report():
    global trade_log

    if not trade_log and os.path.exists(TRADE_LOG_FILE):
        try:
            csv_df = pd.read_csv(TRADE_LOG_FILE)
            trade_log = csv_df.to_dict('records')
        except:
            pass

    if not trade_log:
        return

    df = pd.DataFrame(trade_log)

    wb = Workbook()
    ws_log = wb.active
    ws_log.title = "Trade Log"

    banner_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    banner_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    headers = ["Date", "Symbol", "Side", "Entry Time", "Entry Price",
               "Exit Time", "Exit Price", "SL", "Target 1:2", "Target 1:3",
               "Exit Reason", "Move (Pts)", "Move (%)", "Paper P&L (₹)"]

    for col, header in enumerate(headers, 1):
        cell = ws_log.cell(row=1, column=col, value=header)
        cell.fill = banner_fill
        cell.font = banner_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        values = [
            row_data.date, row_data.symbol, row_data.side, row_data.entry_time,
            row_data.entry_price, row_data.exit_time, row_data.exit_price, row_data.sl,
            getattr(row_data, 'target1_1_2', None), getattr(row_data, 'target2_1_3', None),
            row_data.exit_reason, row_data.move_captured_points,
            getattr(row_data, 'move_captured__', None), row_data.paper_pnl
        ]
        for col, value in enumerate(values, 1):
            cell = ws_log.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if col in [5, 7, 8, 9, 10]:
                cell.number_format = '0.00'
            elif col in [12, 14]:
                cell.number_format = '0.00'
                if value is not None and value > 0:
                    cell.fill = green_fill
                elif value is not None and value < 0:
                    cell.fill = red_fill

    for col in range(1, len(headers) + 1):
        ws_log.column_dimensions[get_column_letter(col)].width = 15

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    total_trades = len(df)
    wins = len(df[df['paper_pnl'] > 0]) if 'paper_pnl' in df.columns else 0
    total_pnl = df['paper_pnl'].sum() if 'paper_pnl' in df.columns else 0
    win_rate = (wins / total_trades * 100) if total_trades > 0 else 0

    ws_summary['A1'] = "PAPER TRADING REPORT (OI + F&O + INDICATOR)"
    ws_summary['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws_summary['A3'] = f"Total Trades: {total_trades}"
    ws_summary['A4'] = f"Win Rate: {win_rate:.1f}%"
    ws_summary['A5'] = f"Total P&L: ₹{round(total_pnl, 2)}"

    wb.save(EXCEL_REPORT_FILE)
    print(f"📊 Report saved: {EXCEL_REPORT_FILE}")

    global _last_report_stats
    _last_report_stats = {
        "total_trades": total_trades,
        "total_pnl": round(total_pnl, 2),
        "win_rate": f"{win_rate:.1f}%"
    }

def send_excel_report_to_telegram():
    if not os.path.exists(EXCEL_REPORT_FILE):
        return
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
        with open(EXCEL_REPORT_FILE, 'rb') as f:
            files = {"document": (EXCEL_REPORT_FILE, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            caption = "📊 Paper Trading Report (OI + F&O + Indicator)"
            if _last_report_stats:
                caption += f"\nTrades: {_last_report_stats.get('total_trades')} | P&L: ₹{_last_report_stats.get('total_pnl')}"
            data = {"chat_id": TELEGRAM_CHAT_ID, "caption": caption}
            requests.post(url, data=data, files=files, timeout=30)
    except Exception as e:
        print(f"❌ Error sending Excel: {e}")

# ==================== MAIN ====================
if __name__ == "__main__":
    import sys
    run_once = os.getenv("RUN_ONCE", "false").lower() == "true"

    if run_once:
        print("🔹 RUN_ONCE mode (OI + F&O + Indicator)")
        now = datetime.now(IST)
        current = now.strftime("%H:%M")

        test_msg = f"🧪 PAPER TRADER (OI+F&O) TEST\\nTime: {now.strftime('%H:%M IST')}\\nInside window: {is_within_alert_window(now)}"
        send_telegram(test_msg, parse_mode=None)

        if now.weekday() >= 5 or not ("09:15" <= current <= "15:30"):
            print("Outside market hours. Exiting.")
            sys.exit(0)

        run_paper_trader()
        sys.exit(0)
    else:
        run_paper_trader()