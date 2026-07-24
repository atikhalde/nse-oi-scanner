"""End-of-day / on-demand paper-test report (xlsx) + telegram summary text.
Supports the setup-aware exits v2 trade format (legs) and gracefully renders
older single-exit trade dicts too."""
import os
from datetime import datetime
import costs as Costs
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill("solid", fgColor="0F172A")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
BUY_FILL = PatternFill("solid", fgColor="00783C")
SELL_FILL = PatternFill("solid", fgColor="9B2D2D")
SKIP_FILL = PatternFill("solid", fgColor="F5F2E8")
W_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(color="0F172A", bold=True, size=14)
BASE_FONT = Font(size=10)
THIN = Border(*[Side(style="thin", color="D0D7E2")] * 4)
GREEN, RED = Font(color="00783C", bold=True, size=10), Font(color="9B2D2D", bold=True, size=10)
GREY = Font(color="6B7280", size=9)

COLS = ["Date", "Symbol", "Side", "Signal", "Setup", "SpurtRank", "Entry time", "Entry ₹", "Qty",
        "Capital ₹", "SL ₹", "Exit", "Exit time", "Avg exit ₹", "P&L ₹", "P&L %", "R", "Costs+Slip ₹", "Net P&L ₹"]

RULES_NOTE = ("setup-aware exits v3 · MIS 5×: ₹10,000 margin → ₹50k notional/trade · structure SL ∓0.02% · "
              "risk ≤ ₹1,000/trade (₹900 planned, qty shrunk; expensive stocks risk less) · NO fixed targets — "
              "100% rides structure-swing trail after +1R to sq-off 15:20 · 90/290 blocked · "
              "B2 quality: only EX1-EX8 + NORMAL · EX9+ blocked (GHOST-shadowed EOD in both SL styles — evidence "
              "build for the release decision, never traded) · EX1/EX2 from 09:45 · " + Costs.NOTE)


def exit_path(tr):
    if tr.get("exit_text"):                       # v2 engine: pre-built from legs
        return tr["exit_text"]
    if tr.get("leg1_why") == "SL":
        return f"SL {tr['leg2_time']}"
    if tr.get("leg1_why") and str(tr.get("leg1_why")).startswith("TP1"):
        e2 = "EOD 15:20" if tr.get("leg2_why") == "EOD15:20" else f"{tr.get('leg2_why')} {tr.get('leg2_time')}"
        return f"50% TP1@1:2 {tr['leg1_time']} · 50% {e2}"
    if tr.get("leg2_why") == "OPEN":
        return "OPEN"
    return "EOD 15:20" if tr.get("leg2_why") == "EOD15:20" else str(tr.get("leg2_why"))


def _avg_exit(tr):
    """Weighted average exit across legs (v2) or two-leg fallback (v1)."""
    if tr.get("legs"):
        qs = sum(q for _l, q, _p, _t in tr["legs"])
        return round(sum(p * q for _l, q, p, _t in tr["legs"]) / (qs or 1), 2)
    avg = tr["leg1_px"] if tr.get("leg1_px") is not None else tr["leg2_px"]
    return round((avg + tr["leg2_px"]) / 2, 2)


def _gnet(rows, mode):
    vals = [g[f"net_{mode}"] for g in rows if g.get(f"net_{mode}") is not None]
    return round(sum(vals), 0), len(vals)


def build(trades, date_lbl, gate_meta, out_path, skipped=None, rules_note=None, ghosts=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Paper test"
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value=f"PAPER TEST — MASTER SCANNER + TOP-30 OI-SPURT GATE — {date_lbl}").font = TITLE_FONT
    note = rules_note or RULES_NOTE
    src = gate_meta.get("source") if isinstance(gate_meta, dict) else None
    ws.cell(row=2, column=1, value=f"OI gate: {gate_meta.get('status')} ({src}) · {note}"
            if isinstance(gate_meta, dict) and "status" in gate_meta else str(gate_meta)).font = Font(size=9)
    hr = 4
    for j, col in enumerate(COLS, 1):
        c = ws.cell(row=hr, column=j, value=col)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, THIN
        ws.column_dimensions[get_column_letter(j)].width = max(11, min(len(col) + 4, 30))
    ws.column_dimensions["L"].width = 34
    r = hr + 1
    tot = 0.0
    for tr in trades:
        sign = 1 if tr["side"] == "BUY" else -1
        avg_ex = _avg_exit(tr)
        pnl_pct = sign * (avg_ex - tr["entry"]) / tr["entry"] * 100
        tot += tr["pnl"]
        cst = Costs.trade_costs(tr)
        vals = [date_lbl, tr["symbol"], tr["side"], tr["signal"], tr.get("setup", "-"),
                tr.get("gate_rank", tr.get("spurt_rank", "-")), tr["time"], tr["entry"],
                tr["qty"], tr.get("capital", round(tr["qty"] * tr["entry"], 0)), tr["sl"],
                exit_path(tr), tr.get("leg2_time"), avg_ex, tr["pnl"], round(pnl_pct, 2), tr["r_total"],
                cst["drag"], cst["net"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font, c.border = BASE_FONT, THIN
            if COLS[j - 1] == "Side":
                c.fill = BUY_FILL if v == "BUY" else SELL_FILL
                c.font = W_FONT
            if COLS[j - 1] in ("P&L ₹", "R", "P&L %", "Net P&L ₹"):
                c.font = GREEN if (v or 0) > 0 else RED
            if COLS[j - 1] in ("Capital ₹", "P&L ₹", "Costs+Slip ₹", "Net P&L ₹"):
                c.number_format = "#,##0"
        r += 1
    wins = sum(1 for t in trades if t["pnl"] > 0)
    c = ws.cell(row=r, column=11,
                value=f"TOTAL ({len(trades)} trades · {wins} wins · {wins * 100 // max(1, len(trades))}%)")
    c.font = Font(bold=True)
    c = ws.cell(row=r, column=15, value=round(tot, 0))
    c.font = GREEN if tot > 0 else RED
    c.number_format = "#,##0"
    rsum = sum(t.get("r_total", 0) for t in trades)
    c = ws.cell(row=r, column=17, value=round(rsum, 2))
    c.font = GREEN if rsum > 0 else RED
    costs_tot = sum(Costs.trade_costs(t)["drag"] for t in trades)
    c = ws.cell(row=r, column=18, value=round(costs_tot, 0))
    c.font = RED; c.number_format = "#,##0"
    c = ws.cell(row=r, column=19, value=round(tot - costs_tot, 0))
    c.font = GREEN if tot - costs_tot > 0 else RED; c.number_format = "#,##0"

    # --- optional disclosure block: signals that were NOT traded
    if skipped:
        r += 2
        c = ws.cell(row=r, column=1,
                    value=f"SKIPPED — NOT TRADED ({sum(len(v) for v in skipped.values())})")
        c.font = Font(color="0F172A", bold=True, size=10)
        for why, items in skipped.items():
            r += 1
            c = ws.cell(row=r, column=1, value=f"{why} ({len(items)})")
            c.font = Font(bold=True, size=9)
            c.fill = SKIP_FILL
            r += 1
            for j, col in enumerate(["Symbol", "Side", "Signal", "Entry time", "Signal ₹"], 1):
                c = ws.cell(row=r, column=j, value=col)
                c.fill, c.font, c.border = SKIP_FILL, Font(bold=True, size=9), THIN
            r += 1
            for it in items:
                for j, v in enumerate(it, 1):
                    c = ws.cell(row=r, column=j, value=v)
                    c.font, c.border = BASE_FONT, THIN
                    if j == 2:
                        c.fill = BUY_FILL if v == "BUY" else SELL_FILL
                        c.font = W_FONT
                r += 1
        ws.cell(row=r + 1, column=1, value="All rows above are real chart signals; skipped only by the "
                                            "trade-selection/flat rule — signal fidelity itself is untouched.").font = GREY
        r += 2

    # --- optional GHOST shadow block: weak-EX signals blocked live, paper-evaluated
    #     after the close in BOTH stop styles. Evidence for the release decision.
    if ghosts:
        tot_s, n_s = _gnet(ghosts, "structure")
        tot_p, n_p = _gnet(ghosts, "prevbar")
        c = ws.cell(row=r, column=1,
                    value=f"👻 GHOST SHADOW — weak-EX signals blocked live, paper-only ({len(ghosts)}) · "
                          f"NO exposure, no intraday alerts — release-decision evidence")
        c.font = Font(color="0F172A", bold=True, size=10)
        r += 1
        c = ws.cell(row=r, column=1,
                    value=f"had they been traded → structure-SL NET ₹{tot_s:+,.0f} ({n_s}) · "
                          f"prev-candle-SL NET ₹{tot_p:+,.0f} ({n_p}) — costs+slippage included")
        c.font = Font(bold=True, size=9)
        r += 1
        has_mv = any("mv" in g for g in ghosts)
        if has_mv:
            gate_rows = [g for g in ghosts if g.get("rank") is not None and g.get("mv")]
            gate_lbl = "spurts-any × movers-20 gate"
        else:
            gate_rows = [g for g in ghosts if g.get("rank") is not None and g["rank"] <= 30]
            gate_lbl = "TOP-30 spurt gate"
        g_s, gn_s = _gnet(gate_rows, "structure")
        g_p, gn_p = _gnet(gate_rows, "prevbar")
        c = ws.cell(row=r, column=1,
                    value=f"… of which pass this model's live entry gate ({gate_lbl}): {len(gate_rows)} signals "
                          f"→ struct NET ₹{g_s:+,.0f} ({gn_s}) · prevbar NET ₹{g_p:+,.0f} ({gn_p})")
        c.font = GREY
        r += 1
        for j, col in enumerate(["Symbol", "Side", "Signal", "Time", "Entry ₹", "Rank", "Grp",
                                 "NET ₹ structure-SL", "Exit (structure)", "NET ₹ prev-candle-SL", "Exit (prevbar)"], 1):
            c = ws.cell(row=r, column=j, value=col)
            c.fill, c.font, c.border = SKIP_FILL, Font(bold=True, size=9), THIN
        r += 1
        for g in ghosts:
            vals = [g["symbol"], g["side"], g["signal"], g["time"], g["entry"],
                    g.get("rank") if g.get("rank") is not None else "-", g.get("grp", ""),
                    g.get("net_structure"), g.get("exit_structure") or "-",
                    g.get("net_prevbar"), g.get("exit_prevbar") or "-"]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=j, value=v)
                c.font, c.border = BASE_FONT, THIN
                if j == 2:
                    c.fill = BUY_FILL if v == "BUY" else SELL_FILL
                    c.font = W_FONT
                if j in (8, 10) and isinstance(v, (int, float)):
                    c.font = GREEN if v > 0 else RED
                    c.number_format = "#,##0"
            r += 1
        ws.cell(row=r, column=1, value="Ghost rows are shadows only — blocked weak-EX chart signals replayed after "
                                       "the close; nothing was traded from this table. Grp R = release candidate "
                                       "(user spec, prev-candle SL) · EX9 = BUY-EX9, blocked either way.").font = GREY
    wb.save(out_path)
    return out_path


def summary_text(trades, date_lbl, gate_meta, ghosts=None):
    wins = [t for t in trades if t["pnl"] > 0]
    tot = sum(t["pnl"] for t in trades)
    rsum = sum(t["r_total"] for t in trades)
    st = gate_meta.get("status") if isinstance(gate_meta, dict) else gate_meta
    src = gate_meta.get("source") if isinstance(gate_meta, dict) else ""
    costs_tot = sum(Costs.trade_costs(t)["drag"] for t in trades)
    wins_net = [t for t in trades if Costs.trade_costs(t)["net"] > 0]
    lines = [f"📊 <b>PAPER TEST {date_lbl} — EOD</b>",
             f"Gate: {st} ({src}) · exits v2 · costs+slippage included (₹20-brokerage + STT + txn + GST + fill haircut)",
             f"Trades: {len(trades)} · Wins (gross/net): {len(wins)}/{len(wins_net)}",
             f"P&L gross: <b>₹{tot:+,.0f}</b> · costs+slip −₹{costs_tot:,.0f} · <b>NET ₹{tot-costs_tot:+,.0f}</b> · {rsum:+.2f}R"]
    for t in trades:
        tag = f"{t.get('setup','')[:4]} " if t.get("setup") else ""
        cn = Costs.trade_costs(t)
        lines.append(f"• {t['symbol']} {t['side']} {tag}{t['signal']} ₹{t['pnl']:+,.0f} (net {cn['net']:+,.0f}, c{cn['drag']:.0f}) {exit_path(t)}")
    if ghosts:
        ts, _ = _gnet(ghosts, "structure")
        tp, _ = _gnet(ghosts, "prevbar")
        lines.append(f"👻 GHOST weak-EX shadow (NOT traded): {len(ghosts)} signals · struct-SL NET ₹{ts:+,.0f} · "
                     f"prevbar-SL NET ₹{tp:+,.0f} — release evidence, session logged")
    return "\n".join(lines)
